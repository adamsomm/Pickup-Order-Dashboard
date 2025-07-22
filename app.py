from flask import Flask, render_template, redirect, request
from openpyxl import load_workbook
from order import PickupOrder
import csv
import os


app = Flask(__name__)

def load_orders_from_excel():
    wb = load_workbook('C:\\Users\\matta\\OneDrive - LC\\Lakelands Pickup Order Submission.xlsx')
    ws = wb.active

    blacklist = get_blacklisted_ids()

    orders = []
    for idx, row_values in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
        unique_id = str(row_values[0]).strip() if row_values[0] is not None else ""
        contact = row_values[5]
        company = row_values[6]
        titan_number = row_values[7]
        details = row_values[8]
        pickup_date = row_values[9]

        if contact is not None and unique_id not in blacklist:
            order = PickupOrder(unique_id, contact, company, details, pickup_date, titan_number)
            order.excel_row = idx
            orders.append(order)

    return orders


def get_blacklisted_ids():
    try:
        with open('blacklist.csv', mode='r', newline='') as file:
            reader = csv.reader(file)
            # Skip header if present
            rows = list(reader)
            if rows and rows[0][0].lower().strip() == 'orderid':
                rows = rows[1:]
            return {row[0].strip() for row in rows if row and row[0].strip()}
    except FileNotFoundError:
        return set()


if not os.path.exists('blacklist.csv'):
    with open('blacklist.csv', mode='w', newline='') as file:
        writer = csv.writer(file)
        writer.writerow(['OrderID'])


@app.route('/')
def dashboard():
    orders = load_orders_from_excel()
    return render_template('dashboard.html', orders=orders)

@app.route('/update-orders', methods=['POST'])
def update_orders():
    remove_ids = request.form.getlist('remove_rows')
    remove_ids = [uid.strip() for uid in remove_ids if uid.strip()]

    if remove_ids:
        existing_blacklist = get_blacklisted_ids()
        new_ids = [uid for uid in remove_ids if uid not in existing_blacklist]

        if new_ids:
            with open('blacklist.csv', mode='a', newline='') as file:
                writer = csv.writer(file)
                for uniqueID in new_ids:
                    writer.writerow([uniqueID])

    return redirect('/')



if __name__ == '__main__':
    app.run(debug=True)
